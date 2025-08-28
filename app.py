"""
Streamlit app for Advanced Cellular Template Processing

Usage:
 - Deploy to Streamlit Community Cloud (or run locally with `streamlit run app.py`)
 - In the sidebar enter your Apify/OpenRouter API key (it should contain 'apify_api_').
 - Validate the key using the "Validate API key" button.
 - After validation, upload your .xlsx or .csv template and images (CSV-case: upload images or a ZIP).
 - Click "Process file", wait for completion, then download the processed file.

Requirements: see requirements.txt (listed below in this message)
"""

import os
import io
import re
import json
import time
import zipfile
import shutil
import base64
import requests
from pathlib import Path
from typing import Optional, Tuple, List

import streamlit as st
import openpyxl
from PIL import Image

# ---------------- CONFIGURATION ----------------
API_BASE = "https://openrouter.apify.actor/api/v1"

# Default models (you can change in UI if desired)
MODEL_SERVICE_DEFAULT = "google/gemini-2.5-flash"
MODEL_GENERIC_DEFAULT = "google/gemini-2.5-flash-lite"

# ---------------- SCHEMAS ----------------
SERVICE_SCHEMA = {
    "nr_arfcn": "number", "nr_band": "number", "nr_pci": "number", "nr_bw": "number",
    "nr5g_rsrp": "number", "nr5g_rsrq": "number", "nr5g_sinr": "number",
    "lte_band": "number", "lte_earfcn": "number", "lte_pci": "number",
    "lte_bw": "number", "lte_rsrp": "number", "lte_rsrq": "number", "lte_sinr": "number"
}

GENERIC_SCHEMAS = {
    "speed_test": {
        "image_type": "speed_test",
        "data": {
            "download_mbps": "number",
            "upload_mbps": "number",
            "ping_ms": "number",
            "jitter_ms": "number"
        }
    },
    "video_test": {
        "image_type": "video_test",
        "data": {
            "max_resolution": "string",
            "load_time_ms": "number",
            "buffering_percentage": "number"
        }
    },
    "voice_call": {
        "image_type": "voice_call",
        "data": {
            "phone_number": "string",
            "call_duration_seconds": "number",
            "call_status": "string",
            "time": "string"
        }
    }
}

# ---------------- GLOBAL HANDLES (reset per run) ----------------
# We'll reinitialize these inside each run to avoid cross-session state issues.
# But keep the names for functions that expect them.
alpha_service = {}
beta_service = {}
gamma_service = {}

alpha_speedtest = {}
beta_speedtest = {}
gamma_speedtest = {}

alpha_video = {}
beta_video = {}
gamma_video = {}

voice_test = {}

extract_text = []
avearge = {}

# ---------------- HELPERS: logging/headers ----------------
def _apify_headers(token: str):
    """Return headers for Apify/OpenRouter calls."""
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "HTTP-Referer": "http://localhost",
        "X-Title": "Advanced Cellular Template Processor"
    }

def log_append(log_placeholder, logs_list, msg: str):
    """Append a message to session logs and refresh the placeholder text area."""
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    logs_list.append(line)
    try:
        log_placeholder.text_area("Logs", "\n".join(logs_list[-1000:]), height=300)
    except Exception:
        # fallback: print
        print(line)

# ---------------- HELPERS: image extraction ----------------
def get_sector_from_col(col_index: int) -> str:
    """Determines sector name from 0-indexed column."""
    if 0 <= col_index < 4:
        return "alpha"
    if 4 <= col_index < 8:
        return "beta"
    if 8 <= col_index < 12:
        return "gamma"
    if 12 <= col_index < 18:
        return "voicetest"
    return "unknown"

def extract_images_from_excel(xlsx_path: str, output_folder: str, log_placeholder, logs) -> List[str]:
    """Extract embedded images from an xlsx workbook and save them to output_folder."""
    log_append(log_placeholder, logs, f"[LOG] Analyzing template file: {xlsx_path}")
    try:
        workbook = openpyxl.load_workbook(xlsx_path)
        sheet = workbook.active
    except Exception as e:
        log_append(log_placeholder, logs, f"[ERROR] Could not open or read the Excel file. {e}")
        return []

    images_with_locations = []
    if not sheet._images:
        log_append(log_placeholder, logs, "[WARN] No images found in the Excel sheet.")
        return []

    for image in sheet._images:
        try:
            row = image.anchor._from.row + 1
            col = image.anchor._from.col
            images_with_locations.append({"image": image, "row": row, "col": col})
        except Exception:
            # fallback if anchor attributes differ
            images_with_locations.append({"image": image, "row": 0, "col": 0})

    sorted_images = sorted(images_with_locations, key=lambda i: (i['row'], i['col']))

    saved_image_paths = []
    sector_counters = {"alpha": 0, "beta": 0, "gamma": 0, "voicetest": 0, "unknown": 0}

    log_append(log_placeholder, logs, f"[LOG] Found {len(sorted_images)} images. Extracting and naming them...")
    for item in sorted_images:
        sector = get_sector_from_col(item['col'])
        sector_counters[sector] += 1
        filename = f"{sector}_image_{sector_counters[sector]}.png"
        output_path = os.path.join(output_folder, filename)

        try:
            img_data = item['image']._data()
            pil_img = Image.open(io.BytesIO(img_data))
            pil_img.save(output_path, 'PNG')
            saved_image_paths.append(output_path)
            loc = ""
            try:
                loc = f"(from cell approx. {openpyxl.utils.get_column_letter(item['col']+1)}{item['row']})"
            except Exception:
                loc = ""
            log_append(log_placeholder, logs, f"  - Saved '{filename}' {loc}")
        except Exception as e:
            log_append(log_placeholder, logs, f"[ERROR] Failed to save image {filename}. Error: {e}")

    return saved_image_paths

def get_images_from_uploads(uploaded_files, output_folder, log_placeholder, logs) -> List[str]:
    """
    Accepts a list of Streamlit uploaded files. Supports raw images and zip containing images.
    Saves to output_folder and returns list of saved image paths.
    """
    os.makedirs(output_folder, exist_ok=True)
    saved = []
    for uf in uploaded_files:
        fname = uf.name.lower()
        if fname.endswith(".zip"):
            # extract images from zip
            try:
                z = zipfile.ZipFile(io.BytesIO(uf.read()))
                for zi in z.infolist():
                    if zi.filename.lower().endswith((".png", ".jpg", ".jpeg")):
                        out_path = os.path.join(output_folder, os.path.basename(zi.filename))
                        with open(out_path, "wb") as f:
                            f.write(z.read(zi.filename))
                        saved.append(out_path)
                log_append(log_placeholder, logs, f"[LOG] Extracted images from zip '{uf.name}' -> {len(saved)} files")
            except Exception as e:
                log_append(log_placeholder, logs, f"[ERROR] Failed to extract zip '{uf.name}': {e}")
        else:
            # assume image file
            try:
                out_path = os.path.join(output_folder, uf.name)
                with open(out_path, "wb") as f:
                    f.write(uf.read())
                saved.append(out_path)
            except Exception as e:
                log_append(log_placeholder, logs, f"[ERROR] Could not save upload '{uf.name}': {e}")
    log_append(log_placeholder, logs, f"[LOG] Saved {len(saved)} uploaded images to temp folder")
    return sorted(saved)

# ---------------- API CALLS: analyzers ----------------
def process_service_images(token: str, image1_path: str, image2_path: str, model_name: str, log_placeholder, logs) -> Optional[dict]:
    """Specialized service extraction that analyzes two images together and returns SERVICE_SCHEMA JSON."""
    sector = Path(image1_path).stem.split('_')[0]
    log_append(log_placeholder, logs, f"[LOG] Starting specialized service data extraction for '{sector}' sector.")
    log_append(log_placeholder, logs, f"[LOG] Using model: {model_name}")
    try:
        with open(image1_path, "rb") as f:
            b64_img1 = base64.b64encode(f.read()).decode('utf-8')
        with open(image2_path, "rb") as f:
            b64_img2 = base64.b64encode(f.read()).decode('utf-8')
    except Exception as e:
        log_append(log_placeholder, logs, f"[ERROR] Could not read or encode service images: {e}")
        return None

    prompt = f"""
You are a hyper-specialized AI for cellular network engineering data analysis. Analyze both provided service-mode screenshots carefully.
Return exactly one JSON object matching schema below. Use JSON null for missing fields.

SCHEMA:
{json.dumps(SERVICE_SCHEMA, indent=2)}
"""

    payload = {
        "model": model_name,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_img1}"}},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_img2}"}},
                ]
            }
        ],
        "response_format": {"type": "json_object"}
    }

    try:
        response = requests.post(url=f"{API_BASE}/chat/completions", headers=_apify_headers(token), data=json.dumps(payload), timeout=120)
        response.raise_for_status()
        content = response.json()['choices'][0]['message']['content']
        result = json.loads(content)
        log_append(log_placeholder, logs, f"[SUCCESS] AI successfully processed service data for '{sector}'.")
        return result
    except Exception as e:
        log_append(log_placeholder, logs, f"[ERROR] API call failed for service images. Error: {e}")
        if 'response' in locals():
            log_append(log_placeholder, logs, f"  Response: {getattr(response, 'text', '')}")
        return None
    finally:
        log_append(log_placeholder, logs, "[LOG] Cooldown: Waiting for 2 seconds...")
        time.sleep(2)

def analyze_generic_image(token: str, image_path: str, model_name: str, log_placeholder, logs) -> Optional[dict]:
    """Classify and extract data from a generic image (speed/video/voice)."""
    image_name = Path(image_path).name
    log_append(log_placeholder, logs, f"[LOG] Starting generic data extraction for '{image_name}'.")
    log_append(log_placeholder, logs, f"[LOG] Using model: {model_name}")
    try:
        with open(image_path, "rb") as f:
            b64_img = base64.b64encode(f.read()).decode('utf-8')
    except Exception as e:
        log_append(log_placeholder, logs, f"[ERROR] Could not read or encode image '{image_name}': {e}")
        return None

    prompt = f"""
You are an expert AI assistant for analyzing cellular network test data. Classify the image as 'speed_test', 'video_test', or 'voice_call'.
Return exactly one JSON object matching the corresponding schema. Use JSON null for missing values.

SCHEMAS:
{json.dumps(GENERIC_SCHEMAS, indent=2)}
"""

    payload = {
        "model": model_name,
        "messages": [
            {"role": "user", "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_img}"}}]}
        ],
        "response_format": {"type": "json_object"}
    }

    try:
        response = requests.post(url=f"{API_BASE}/chat/completions", headers=_apify_headers(token), data=json.dumps(payload), timeout=60)
        response.raise_for_status()
        content = response.json()['choices'][0]['message']['content']
        result = json.loads(content)
        log_append(log_placeholder, logs, f"[SUCCESS] AI successfully processed '{image_name}' as type '{result.get('image_type', 'unknown')}'.")
        return result
    except Exception as e:
        log_append(log_placeholder, logs, f"[ERROR] API call failed for '{image_name}'. Error: {e}")
        if 'response' in locals():
            log_append(log_placeholder, logs, f"  Response: {getattr(response, 'text', '')}")
        return None
    finally:
        log_append(log_placeholder, logs, "[LOG] Cooldown: Waiting for 2 seconds...")
        time.sleep(2)

def analyze_voice_image(token: str, image_path: str, model_name: str, log_placeholder, logs) -> Optional[dict]:
    """Voice-specific analyzer (ensures 'time' field is extracted)."""
    image_name = Path(image_path).name
    log_append(log_placeholder, logs, f"[VOICE] Starting voice-specific extraction for '{image_name}'.")
    try:
        with open(image_path, "rb") as f:
            b64_img = base64.b64encode(f.read()).decode('utf-8')
    except Exception as e:
        log_append(log_placeholder, logs, f"[VOICE ERROR] Could not read/encode image '{image_name}': {e}")
        return None

    prompt = f"""
You are an expert in telecom voice-call screenshot extraction. Extract ONLY the fields in the voice_call schema below, with emphasis on 'time'.
Return exactly one JSON object matching the schema, or null for missing fields.

SCHEMA:
{json.dumps(GENERIC_SCHEMAS['voice_call'], indent=2)}
"""

    payload = {
        "model": model_name,
        "messages": [
            {"role": "user", "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_img}"}}]}
        ],
        "response_format": {"type": "json_object"}
    }

    try:
        response = requests.post(url=f"{API_BASE}/chat/completions", headers=_apify_headers(token), data=json.dumps(payload), timeout=60)
        response.raise_for_status()
        content = response.json()['choices'][0]['message']['content']
        res = json.loads(content)
        log_append(log_placeholder, logs, f"[VOICE SUCCESS] Processed '{image_name}'.")
        return res
    except Exception as e:
        log_append(log_placeholder, logs, f"[VOICE ERROR] API call failed for '{image_name}': {e}")
        if 'response' in locals():
            log_append(log_placeholder, logs, f"  Response: {getattr(response, 'text', '')}")
        return None
    finally:
        log_append(log_placeholder, logs, "[VOICE] Cooldown: Waiting for 2 seconds...")
        time.sleep(2)

# ---------------- EVALUATION (careful re-check) ----------------
def evaluate_service_images(token: str, image1_path: str, image2_path: str, model_name: str, log_placeholder, logs) -> Optional[dict]:
    """Careful line-by-line re-evaluation of service images."""
    sector = Path(image1_path).stem.split('_')[0]
    log_append(log_placeholder, logs, f"[EVAL] Re-evaluating service images for '{sector}' (careful mode).")
    try:
        with open(image1_path, "rb") as f:
            b64_img1 = base64.b64encode(f.read()).decode('utf-8')
        with open(image2_path, "rb") as f:
            b64_img2 = base64.b64encode(f.read()).decode('utf-8')
    except Exception as e:
        log_append(log_placeholder, logs, f"[EVAL ERROR] Could not read or encode service images: {e}")
        return None

    prompt = f"""
You are an expert cellular network data extraction system. THIS IS A CAREFUL, LINE-BY-LINE EVALUATION.
Return exactly one JSON object matching the SCHEMA below. Use null only where field cannot be found.

SCHEMA:
{json.dumps(SERVICE_SCHEMA, indent=2)}
"""

    payload = {
        "model": model_name,
        "messages": [
            {"role": "user", "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_img1}"}}},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_img2}"}}]
            }
        ],
        "response_format": {"type": "json_object"}
    }

    try:
        response = requests.post(url=f"{API_BASE}/chat/completions", headers=_apify_headers(token), data=json.dumps(payload), timeout=120)
        response.raise_for_status()
        content = response.json()['choices'][0]['message']['content']
        return json.loads(content)
    except Exception as e:
        log_append(log_placeholder, logs, f"[EVAL ERROR] Service evaluation API failed: {e}")
        if 'response' in locals():
            log_append(log_placeholder, logs, f"  Response: {getattr(response, 'text', '')}")
        return None
    finally:
        log_append(log_placeholder, logs, "[EVAL] Cooldown: Waiting for 2 seconds...")
        time.sleep(2)

def evaluate_generic_image(token: str, image_path: str, model_name: str, log_placeholder, logs) -> Optional[dict]:
    """Careful evaluation for a single generic image."""
    image_name = Path(image_path).name
    log_append(log_placeholder, logs, f"[EVAL] Re-evaluating image '{image_name}' (careful mode).")
    try:
        with open(image_path, "rb") as f:
            b64_img = base64.b64encode(f.read()).decode('utf-8')
    except Exception as e:
        log_append(log_placeholder, logs, f"[EVAL ERROR] Could not read or encode image '{image_name}': {e}")
        return None

    prompt = f"""
You are an expert cellular network data extraction system. THIS IS A CAREFUL, LINE-BY-LINE EVALUATION.
Classify the image and return exactly one JSON object matching the schema. Use null only when necessary.

SCHEMAS:
{json.dumps(GENERIC_SCHEMAS, indent=2)}
"""

    payload = {
        "model": model_name,
        "messages": [
            {"role": "user", "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_img}"}}]}
        ],
        "response_format": {"type": "json_object"}
    }

    try:
        response = requests.post(url=f"{API_BASE}/chat/completions", headers=_apify_headers(token), data=json.dumps(payload), timeout=90)
        response.raise_for_status()
        content = response.json()['choices'][0]['message']['content']
        return json.loads(content)
    except Exception as e:
        log_append(log_placeholder, logs, f"[EVAL ERROR] Generic evaluation API failed for '{image_name}': {e}")
        if 'response' in locals():
            log_append(log_placeholder, logs, f"  Response: {getattr(response, 'text', '')}")
        return None
    finally:
        log_append(log_placeholder, logs, "[EVAL] Cooldown: Waiting for 2 seconds...")
        time.sleep(2)

# ---------------- Expression resolution helpers ----------------
key_pattern = re.compile(r"\[['\"]([^'\"]+)['\"]\]")

def _normalize_name(n: str) -> str:
    return re.sub(r'[^0-9a-zA-Z]', '', n).lower()

def resolve_expression_with_vars(expr: str, allowed_vars: dict):
    """
    Attempts to resolve expression like alpha_speedtest['alpha_image_6']['download_mbps']
    Works case-insensitively for base and keys.
    Returns resolved value or None.
    """
    expr = expr.strip()
    m = re.match(r"^([A-Za-z_]\w*)(.*)$", expr)
    if not m:
        return None
    base_raw = m.group(1)
    rest = m.group(2) or ""

    norm_map = { _normalize_name(k): k for k in allowed_vars.keys() }
    base_norm = _normalize_name(base_raw)
    base_key = norm_map.get(base_norm)
    if not base_key:
        # direct case-insensitive fallback
        for k in allowed_vars.keys():
            if k.lower() == base_raw.lower():
                base_key = k
                break
    if not base_key:
        return None

    obj = allowed_vars[base_key]
    if rest.strip() == "":
        return obj

    keys = key_pattern.findall(rest)
    if not keys:
        return None

    try:
        for k in keys:
            if not isinstance(obj, dict):
                return None
            # direct key
            if k in obj:
                obj = obj[k]
                continue
            # case-insensitive match
            found = None
            for real_k in obj.keys():
                if real_k.lower() == k.lower() or _normalize_name(real_k) == _normalize_name(k):
                    found = real_k
                    break
            if found:
                obj = obj[found]
            else:
                return None
        return obj
    except Exception:
        return None

def set_nested_value_case_insensitive(target_dict: dict, keys: list, value):
    """
    Set nested value into dictionary following case-insensitive / normalized keys.
    keys is list of key strings; will try to match existing keys case-insensitively,
    otherwise will create keys using the provided key names.
    """
    cur = target_dict
    for i, k in enumerate(keys):
        if i == len(keys) - 1:
            if isinstance(cur, dict):
                found = None
                for real_k in list(cur.keys()):
                    if real_k.lower() == k.lower() or _normalize_name(real_k) == _normalize_name(k):
                        found = real_k
                        break
                if found:
                    cur[found] = value
                else:
                    cur[k] = value
            return True
        else:
            found = None
            if isinstance(cur, dict):
                for real_k in list(cur.keys()):
                    if real_k.lower() == k.lower() or _normalize_name(real_k) == _normalize_name(k):
                        found = real_k
                        break
            if found:
                if not isinstance(cur[found], dict):
                    cur[found] = {}
                cur = cur[found]
            else:
                cur[k] = {}
                cur = cur[k]
    return True

def ask_model_for_expression_value(token: str, var_name: str, var_obj, expression: str, model_name: str, log_placeholder, logs):
    """
    Ask the model to extract the value of `expression` from the provided JSON `var_obj`.
    Returns Python value or None.
    """
    try:
        var_json = json.dumps(var_obj, indent=2)
    except Exception:
        var_json = json.dumps(str(var_obj))

    prompt = f"""
You are an exacting assistant. You are given a JSON variable named "{var_name}" with the following content:

{var_json}

Given the expression:
{expression}

Using ONLY the provided JSON variable, evaluate the expression and return exactly one JSON object like:
{{ "value": <value> }}

Where <value> is the exact value (number or string) if present, or null if not present. Return only the JSON object and nothing else.
"""
    payload = {
        "model": model_name,
        "messages": [
            {"role":"user","content":[{"type":"text","text":prompt}]}
        ],
        "response_format": {"type":"json_object"}
    }

    try:
        response = requests.post(url=f"{API_BASE}/chat/completions", headers=_apify_headers(token), data=json.dumps(payload), timeout=30)
        response.raise_for_status()
        content = response.json()['choices'][0]['message']['content']
        parsed = json.loads(content)
        return parsed.get("value", None)
    except Exception as e:
        log_append(log_placeholder, logs, f"[ASK-MODEL] Failed to get value from model for expression {expression}: {e}")
        if 'response' in locals():
            log_append(log_placeholder, logs, f"  Response: {getattr(response, 'text', '')}")
        return None

# ---------------- MAIN PROCESSING FUNCTION (streamlit-friendly) ----------------
def process_file_streamlit(user_file_path: str,
                           token: str,
                           temp_dir: str,
                           image_uploads: List[io.BytesIO],
                           logs: list,
                           text_area_placeholder,
                           model_service=MODEL_SERVICE_DEFAULT,
                           model_generic=MODEL_GENERIC_DEFAULT) -> Optional[str]:
    """
    Processes the file (xlsx or csv) and returns path to updated file (inside temp_dir).
    This function mirrors your original CLI flow but allows Streamlit-friendly logging via placeholders.
    """
    # We'll use local copies of the global variables and update them in place
    global alpha_service, beta_service, gamma_service
    global alpha_speedtest, beta_speedtest, gamma_speedtest
    global alpha_video, beta_video, gamma_video
    global voice_test, extract_text, avearge

    # reset for this run
    alpha_service = {}
    beta_service = {}
    gamma_service = {}

    alpha_speedtest = {}
    beta_speedtest = {}
    gamma_speedtest = {}

    alpha_video = {}
    beta_video = {}
    gamma_video = {}

    voice_test = {}
    extract_text = []
    avearge = {}

    # prepare temp directories
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir, exist_ok=True)
    images_temp = os.path.join(temp_dir, "images")
    os.makedirs(images_temp, exist_ok=True)

    # copy or save the uploaded file into temp dir
    try:
        local_template = os.path.join(temp_dir, os.path.basename(user_file_path))
        # if user_file_path is a path-like streamlit saved path, it's already on disk,
        # but in many deployments we get the temp path. We'll detect existence:
        if os.path.exists(user_file_path):
            shutil.copy(user_file_path, local_template)
        else:
            # If user_file_path is a file-like object name, user might have passed path string earlier
            # Fall back to raising error
            raise FileNotFoundError(f"File not found: {user_file_path}")
    except Exception as e:
        log_append(text_area_placeholder, logs, f"[ERROR] Could not copy input file to temp dir: {e}")
        return None

    # 1) extract images (if xlsx) else get images from uploaded set (CSV case)
    path_obj = Path(local_template)
    image_paths = []
    if path_obj.suffix.lower() == ".xlsx":
        image_paths = extract_images_from_excel(local_template, images_temp, text_area_placeholder, logs)
    elif path_obj.suffix.lower() == ".csv":
        # user should have uploaded images in image_uploads (list of UploadedFile)
        if not image_uploads:
            log_append(text_area_placeholder, logs, "[ERROR] CSV input provided but no images were uploaded.")
            return None
        image_paths = get_images_from_uploads(image_uploads, images_temp, text_area_placeholder, logs)
    else:
        log_append(text_area_placeholder, logs, "[ERROR] Unsupported file type. Upload .xlsx or .csv.")
        return None

    if not image_paths:
        log_append(text_area_placeholder, logs, "[ERROR] No images found or extracted. Exiting.")
        return None

    # Group images by sector
    images_by_sector = {"alpha": [], "beta": [], "gamma": [], "voicetest": [], "unknown": []}
    for p in image_paths:
        sector = Path(p).stem.split('_')[0]
        if sector in images_by_sector:
            images_by_sector[sector].append(p)
        else:
            images_by_sector["unknown"].append(p)

    # --- MAIN PROCESSING LOOP ---
    log_append(text_area_placeholder, logs, "[LOG] Beginning main processing loop over sectors.")
    for sector in ["alpha", "beta", "gamma"]:
        log_append(text_area_placeholder, logs, f"\n--- Processing Sector: {sector.upper()} ---")
        sector_images = images_by_sector[sector]

        # Service images: image_1 and image_2
        img1 = next((p for p in sector_images if Path(p).stem.endswith("_image_1")), None)
        img2 = next((p for p in sector_images if Path(p).stem.endswith("_image_2")), None)
        if img1 and img2:
            service_data = process_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
            if service_data:
                if sector == "alpha":
                    alpha_service = service_data
                elif sector == "beta":
                    beta_service = service_data
                elif sector == "gamma":
                    gamma_service = service_data
        else:
            log_append(text_area_placeholder, logs, f"[WARN] Could not find both image_1 and image_2 for sector '{sector}'.")

        # Other images for sector (speed/video/voice)
        other_images = [p for p in sector_images if not (Path(p).stem.endswith("_image_1") or Path(p).stem.endswith("_image_2"))]
        for img_path in other_images:
            result = analyze_generic_image(token, img_path, model_generic, text_area_placeholder, logs)
            if result and 'image_type' in result:
                image_name = Path(img_path).stem
                if result['image_type'] == 'speed_test':
                    if sector == "alpha":
                        alpha_speedtest[image_name] = result['data']
                    elif sector == "beta":
                        beta_speedtest[image_name] = result['data']
                    elif sector == "gamma":
                        gamma_speedtest[image_name] = result['data']
                elif result['image_type'] == 'video_test':
                    if sector == "alpha":
                        alpha_video[image_name] = result['data']
                    elif sector == "beta":
                        beta_video[image_name] = result['data']
                    elif sector == "gamma":
                        gamma_video[image_name] = result['data']
                elif result['image_type'] == 'voice_call':
                    # handle if voice_call appears unexpectedly
                    voice_test[image_name] = result['data']

    # Process voicetest sector separately using voice-specific analyzer
    if images_by_sector["voicetest"]:
        log_append(text_area_placeholder, logs, "\n--- Processing Sector: VOICETEST ---")
        for img_path in images_by_sector["voicetest"]:
            result = analyze_voice_image(token, img_path, model_generic, text_area_placeholder, logs)
            if result and result.get('image_type') == 'voice_call':
                image_name = Path(img_path).stem
                voice_test[image_name] = result['data']

    # ---------------- EVALUATION PASS & RULE 2 ----------------
    log_append(text_area_placeholder, logs, "\n[LOG] Starting evaluation pass to refill missing/null fields (one retry per item).")
    retried_service_sectors = set()
    retried_images = set()

    def contains_nulls(d):
        if not isinstance(d, dict):
            return False
        for v in d.values():
            if v is None:
                return True
            if isinstance(v, dict) and contains_nulls(v):
                return True
        return False

    # Evaluate service dicts
    for sector in ["alpha", "beta", "gamma"]:
        svc_var = {"alpha": alpha_service, "beta": beta_service, "gamma": gamma_service}[sector]
        if not svc_var:
            img1 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_1")), None)
            img2 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_2")), None)
            if img1 and img2 and sector not in retried_service_sectors:
                log_append(text_area_placeholder, logs, f"[EVAL] Service dict for '{sector}' is empty. Re-evaluating service images.")
                eval_res = evaluate_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
                retried_service_sectors.add(sector)
                if eval_res:
                    if sector == "alpha":
                        alpha_service = eval_res
                    if sector == "beta":
                        beta_service = eval_res
                    if sector == "gamma":
                        gamma_service = eval_res
            continue

        if contains_nulls(svc_var) and sector not in retried_service_sectors:
            img1 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_1")), None)
            img2 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_2")), None)
            if img1 and img2:
                log_append(text_area_placeholder, logs, f"[EVAL] Found nulls in {sector}_service; re-evaluating service images for sector '{sector}'.")
                eval_res = evaluate_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
                retried_service_sectors.add(sector)
                if eval_res:
                    target = {"alpha": alpha_service, "beta": beta_service, "gamma": gamma_service}[sector]
                    for k, v in eval_res.items():
                        if (target.get(k) is None) and v is not None:
                            target[k] = v

    # helper for image retry & merge
    def _retry_image_and_merge(image_name, sector_var_map):
        # image_name is like 'alpha_image_3'
        image_path = os.path.join(images_temp, f"{image_name}.png")
        if not os.path.exists(image_path):
            found = None
            for s_list in images_by_sector.values():
                for p in s_list:
                    if Path(p).stem == image_name:
                        found = p
                        break
                if found:
                    break
            if found:
                image_path = found
            else:
                log_append(text_area_placeholder, logs, f"[EVAL WARN] Image file not found for {image_name}. Skipping.")
                return False
        if image_path in retried_images:
            return False

        is_voice = image_name.startswith("voicetest")
        log_append(text_area_placeholder, logs, f"[EVAL] Attempting normal analyze for {image_name}.")
        if is_voice:
            normal_res = analyze_voice_image(token, image_path, model_generic, text_area_placeholder, logs)
        else:
            normal_res = analyze_generic_image(token, image_path, model_generic, text_area_placeholder, logs)

        retried_images.add(image_path)
        if normal_res and 'image_type' in normal_res:
            sector_var_map.setdefault(image_name, {})
            data = normal_res.get('data', {})
            for k, v in data.items():
                if sector_var_map[image_name].get(k) is None and v is not None:
                    sector_var_map[image_name][k] = v
            return True

        log_append(text_area_placeholder, logs, f"[EVAL] Normal analyze did not yield usable data for {image_name}. Trying careful evaluation.")
        if is_voice:
            eval_res = evaluate_voice_image(token, image_path, model_generic, text_area_placeholder, logs)
        else:
            eval_res = evaluate_generic_image(token, image_path, model_generic, text_area_placeholder, logs)

        if not eval_res or 'image_type' not in eval_res:
            log_append(text_area_placeholder, logs, f"[EVAL] Careful evaluation returned nothing usable for {image_name}.")
            return False

        sector_var_map.setdefault(image_name, {})
        for k, v in eval_res.get('data', {}).items():
            if sector_var_map[image_name].get(k) is None and v is not None:
                sector_var_map[image_name][k] = v
        return True

    sector_maps = [
        ("alpha", alpha_speedtest, alpha_video),
        ("beta", beta_speedtest, beta_video),
        ("gamma", gamma_speedtest, gamma_video),
    ]

    expected_indices = {
        "service": [1, 2],
        "speed": [3, 4, 5, 6, 7],
        "video": [8]
    }

    def missing_service_fields(svc_obj):
        missing = []
        for k in SERVICE_SCHEMA.keys():
            if k not in svc_obj or svc_obj.get(k) is None:
                missing.append(k)
        return missing

    for sector, speed_map, video_map in sector_maps:
        log_append(text_area_placeholder, logs, f"[RULE2] Verifying expected images & schema completeness for sector '{sector}'.")
        svc_var = {"alpha": alpha_service, "beta": beta_service, "gamma": gamma_service}[sector]
        svc_missing = missing_service_fields(svc_var) if svc_var else list(SERVICE_SCHEMA.keys())
        if svc_missing:
            log_append(text_area_placeholder, logs, f"[RULE2] Service for '{sector}' missing fields: {svc_missing}")
            img1 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_1")), None)
            img2 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_2")), None)
            if img1 and img2 and sector not in retried_service_sectors:
                log_append(text_area_placeholder, logs, f"[RULE2] Attempting re-process of service images for '{sector}'.")
                normal_svc = process_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
                retried_service_sectors.add(sector)
                if normal_svc:
                    target = {"alpha": alpha_service, "beta": beta_service, "gamma": gamma_service}[sector]
                    for k, v in normal_svc.items():
                        if target.get(k) is None and v is not None:
                            target[k] = v
                    if missing_service_fields(target):
                        log_append(text_area_placeholder, logs, f"[RULE2] Attempting careful evaluation of service images for '{sector}'.")
                        eval_svc = evaluate_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
                        if eval_svc:
                            for k, v in eval_svc.items():
                                if target.get(k) is None and v is not None:
                                    target[k] = v
            else:
                log_append(text_area_placeholder, logs, f"[RULE2] Cannot re-process service for '{sector}': service images missing or already retried.")

        for idx in expected_indices['speed']:
            name = f"{sector}_image_{idx}"
            if name not in speed_map:
                log_append(text_area_placeholder, logs, f"[RULE2] Missing expected speed image {name}. Looking for the file to process.")
                file_path = next((p for p in images_by_sector[sector] if Path(p).stem == name), None)
                if file_path:
                    log_append(text_area_placeholder, logs, f"[RULE2] Found file for {name}. Processing.")
                    _retry_image_and_merge(name, speed_map)
                else:
                    log_append(text_area_placeholder, logs, f"[RULE2] No file found for expected image {name}.")
            else:
                missing = []
                for k in GENERIC_SCHEMAS['speed_test']['data'].keys():
                    if k not in speed_map[name] or speed_map[name].get(k) is None:
                        missing.append(k)
                if missing:
                    log_append(text_area_placeholder, logs, f"[RULE2] {name} has missing speed fields {missing}. Re-evaluating the image.")
                    _retry_image_and_merge(name, speed_map)

        for idx in expected_indices['video']:
            name = f"{sector}_image_{idx}"
            if name not in video_map:
                file_path = next((p for p in images_by_sector[sector] if Path(p).stem == name), None)
                if file_path:
                    log_append(text_area_placeholder, logs, f"[RULE2] {sector}_video missing entry {name}. Processing file.")
                    _retry_image_and_merge(name, video_map)
                else:
                    log_append(text_area_placeholder, logs, f"[RULE2] No file found for expected video {name}.")
            else:
                missing = []
                for k in GENERIC_SCHEMAS['video_test']['data'].keys():
                    if k not in video_map[name] or video_map[name].get(k) is None:
                        missing.append(k)
                if missing:
                    log_append(text_area_placeholder, logs, f"[RULE2] {name} video has missing fields {missing}. Re-evaluating the image.")
                    _retry_image_and_merge(name, video_map)

    # voicetest check
    log_append(text_area_placeholder, logs, "[RULE2] Verifying voicetest images and schema completeness.")
    for idx in [1, 2, 3]:
        name = f"voicetest_image_{idx}"
        if name not in voice_test:
            file_path = next((p for p in images_by_sector["voicetest"] if Path(p).stem == name), None)
            if file_path:
                log_append(text_area_placeholder, logs, f"[RULE2] Missing voice_test entry {name}. Processing file.")
                _retry_image_and_merge(name, voice_test)
            else:
                log_append(text_area_placeholder, logs, f"[RULE2] No file found for expected voice test image {name}.")
        else:
            missing = []
            for k in GENERIC_SCHEMAS['voice_call']['data'].keys():
                if k not in voice_test[name] or voice_test[name].get(k) is None:
                    missing.append(k)
            if missing:
                log_append(text_area_placeholder, logs, f"[RULE2] {name} has missing voice fields {missing}. Re-evaluating the image.")
                _retry_image_and_merge(name, voice_test)

    log_append(text_area_placeholder, logs, "[LOG] Rule 2 verification complete.\n[LOG] Evaluation pass complete.\n")

    # ---------- compute avearge ----------
    def _to_number(v):
        try:
            if v is None:
                return None
            if isinstance(v, bool):
                return None
            return float(v)
        except Exception:
            return None

    def _compute_speed_averages(speed_map):
        metrics = {"download_mbps": [], "upload_mbps": [], "ping_ms": []}
        for entry in speed_map.values():
            if not isinstance(entry, dict):
                continue
            for m in metrics.keys():
                val = _to_number(entry.get(m))
                if val is not None:
                    metrics[m].append(val)
        result = {}
        for m, vals in metrics.items():
            if vals:
                result[m] = sum(vals) / len(vals)
            else:
                result[m] = None
        return result

    avearge = {
        "avearge_alpha_speedtest": _compute_speed_averages(alpha_speedtest),
        "avearge_beta_speedtest": _compute_speed_averages(beta_speedtest),
        "avearge_gamma_speedtest": _compute_speed_averages(gamma_speedtest),
    }

    # -------------------- INITIAL MAPPING: extract bold+red expressions and replace (STRICT) --------------------
    log_append(text_area_placeholder, logs, "[LOG] Scanning workbook for BOLD + RED expressions (columns A..P ONLY) and replacing them with values.")
    try:
        wb_edit = openpyxl.load_workbook(local_template)
        sheet_edit = wb_edit.active

        cells_to_process = []

        def _font_is_strict_red(font):
            if not font:
                return False
            if not getattr(font, "bold", False):
                return False
            col = getattr(font, "color", None)
            if col is None:
                return False
            rgb = getattr(col, "rgb", None)
            if not rgb:
                return False
            up = str(rgb).upper()
            last6 = up[-6:]
            return last6 == "FF0000"

        def _normalize_expr(raw):
            s = raw.strip()
            if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
                s = s[1:-1].strip()
            return s

        for row in sheet_edit.iter_rows(min_row=1, max_row=sheet_edit.max_row, min_col=1, max_col=16):
            for cell in row:
                val = cell.value
                if not val or not isinstance(val, str):
                    continue
                font = cell.font
                if not font:
                    continue
                if _font_is_strict_red(font):
                    expr = _normalize_expr(val)
                    if expr:
                        extract_text.append(expr)
                        cells_to_process.append((cell, expr))

        allowed_vars = {
            "alpha_service": alpha_service,
            "beta_service": beta_service,
            "gamma_service": gamma_service,
            "alpha_speedtest": alpha_speedtest,
            "beta_speedtest": beta_speedtest,
            "gamma_speedtest": gamma_speedtest,
            "alpha_video": alpha_video,
            "beta_video": beta_video,
            "gamma_video": gamma_video,
            "voice_test": voice_test,
            "avearge": avearge,
        }

        def _to_number_convert(v):
            try:
                if v is None:
                    return None
                if isinstance(v, (int, float)):
                    return v
                if isinstance(v, bool):
                    return None
                s = str(v).strip()
                s_clean = s.replace(',', '')
                if s_clean == "":
                    return None
                if re.fullmatch(r"[-+]?\d+", s_clean):
                    return int(s_clean)
                if re.fullmatch(r"[-+]?\d*\.\d+", s_clean):
                    return float(s_clean)
                return None
            except Exception:
                return None

        # initial replacements
        for cell_obj, expr in cells_to_process:
            resolved = resolve_expression_with_vars(expr, allowed_vars)
            if resolved is None:
                cell_obj.value = "NULL"
            else:
                if isinstance(resolved, str):
                    num = _to_number_convert(resolved)
                    if num is not None:
                        cell_obj.value = num
                    else:
                        cell_obj.value = resolved
                elif isinstance(resolved, (int, float)):
                    cell_obj.value = resolved
                elif isinstance(resolved, (dict, list)):
                    try:
                        cell_obj.value = json.dumps(resolved)
                    except Exception:
                        cell_obj.value = str(resolved)
                else:
                    cell_obj.value = str(resolved)

        wb_edit.save(local_template)
        log_append(text_area_placeholder, logs, f"[LOG] Workbook updated and saved to: {local_template}")
    except Exception as e:
        log_append(text_area_placeholder, logs, f"[ERROR] Failed to edit/save workbook with extracted text replacements: {e}")

    # -------------------- RULE 3: remap remaining NULL cells using strict AI re-checks --------------------
    log_append(text_area_placeholder, logs, "[LOG] Running Rule 3: remap any bold+red expressions that remained NULL using careful AI re-checks.")
    try:
        wb_r3 = openpyxl.load_workbook(local_template)
        sheet_r3 = wb_r3.active

        allowed_vars = {
            "alpha_service": alpha_service,
            "beta_service": beta_service,
            "gamma_service": gamma_service,
            "alpha_speedtest": alpha_speedtest,
            "beta_speedtest": beta_speedtest,
            "gamma_speedtest": gamma_speedtest,
            "alpha_video": alpha_video,
            "beta_video": beta_video,
            "gamma_video": gamma_video,
            "voice_test": voice_test,
            "avearge": avearge,
        }

        problematic_cells = []
        for row in sheet_r3.iter_rows(min_row=1, max_row=sheet_r3.max_row, min_col=1, max_col=16):
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                if val.strip().upper() != "NULL":
                    continue
                font = cell.font
                if font and _font_is_strict_red(font):
                    problematic_cells.append(cell)

        remapped_count = 0
        for cell in problematic_cells:
            # try to find candidate expression in extract_text that matches allowed var
            candidate = None
            for expr in extract_text:
                mm = re.match(r"^([A-Za-z_]\w*)", expr.strip())
                if not mm:
                    continue
                base_raw = mm.group(1)
                if _normalize_name(base_raw) in { _normalize_name(k) for k in allowed_vars.keys() }:
                    # prefer expressions that contain the target cell coordinate? Not available, so just pick first unresolved
                    if resolve_expression_with_vars(expr, allowed_vars) is None:
                        candidate = expr
                        break
            if not candidate:
                continue

            expr = candidate
            log_append(text_area_placeholder, logs, f"[RULE3] Attempting strict re-map for expression '{expr}' for cell {cell.coordinate}.")
            m = re.match(r"^([A-Za-z_]\w*)(.*)$", expr)
            if not m:
                log_append(text_area_placeholder, logs, f"[RULE3] Could not parse expression '{expr}'. Skipping.")
                continue
            base_raw = m.group(1)
            rest = m.group(2) or ""
            norm_map = { _normalize_name(k): k for k in allowed_vars.keys() }
            base_key = norm_map.get(_normalize_name(base_raw))
            if not base_key:
                log_append(text_area_placeholder, logs, f"[RULE3] Base variable '{base_raw}' not in allowed_vars. Skipping.")
                continue

            # If service variable, try re-eval service images first
            if base_key in ("alpha_service", "beta_service", "gamma_service"):
                sector = base_key.split('_')[0]
                img1 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_1")), None)
                img2 = next((p for p in images_by_sector[sector] if Path(p).stem.endswith("_image_2")), None)
                if img1 and img2:
                    log_append(text_area_placeholder, logs, f"[RULE3] Re-evaluating service images for sector '{sector}' (strict).")
                    svc_eval = evaluate_service_images(token, img1, img2, model_service, text_area_placeholder, logs)
                    if svc_eval:
                        allowed_vars[base_key].update(svc_eval)
                        resolved_after = resolve_expression_with_vars(expr, allowed_vars)
                        if resolved_after is not None:
                            keys = key_pattern.findall(m.group(2))
                            set_nested_value_case_insensitive(allowed_vars[base_key], keys, resolved_after)
                            if isinstance(resolved_after, (int, float)):
                                cell.value = resolved_after
                            elif isinstance(resolved_after, str):
                                cell.value = resolved_after
                            else:
                                try:
                                    cell.value = json.dumps(resolved_after)
                                except Exception:
                                    cell.value = str(resolved_after)
                            remapped_count += 1
                            continue
                # ask model using provided JSON var
                log_append(text_area_placeholder, logs, f"[RULE3] Asking model for value of '{expr}' using variable '{base_key}'.")
                value = ask_model_for_expression_value(token, base_key, allowed_vars[base_key], expr, model_generic, text_area_placeholder, logs)
                if value is not None:
                    keys = key_pattern.findall(rest)
                    set_nested_value_case_insensitive(allowed_vars[base_key], keys, value)
                    if isinstance(value, (int, float)):
                        cell.value = value
                    elif isinstance(value, str):
                        cell.value = value
                    else:
                        try:
                            cell.value = json.dumps(value)
                        except Exception:
                            cell.value = str(value)
                    remapped_count += 1
                    continue
                else:
                    log_append(text_area_placeholder, logs, f"[RULE3] Model could not produce value for '{expr}'. Leaving NULL.")
                    continue

            # Non-service variables: first key should be image key like 'gamma_image_6'
            keys = key_pattern.findall(rest)
            if not keys:
                log_append(text_area_placeholder, logs, f"[RULE3] No bracketed keys in '{expr}', cannot remap. Skipping.")
                continue
            image_key = keys[0]
            # find file path for that image_key
            file_path = None
            for lst in images_by_sector.values():
                for p in lst:
                    if Path(p).stem == image_key:
                        file_path = p
                        break
                if file_path:
                    break

            if not file_path:
                log_append(text_area_placeholder, logs, f"[RULE3] Could not find file for image '{image_key}'. Asking model using variable '{base_key}'.")
                value = ask_model_for_expression_value(token, base_key, allowed_vars[base_key], expr, model_generic, text_area_placeholder, logs)
                if value is not None:
                    set_nested_value_case_insensitive(allowed_vars[base_key], keys[1:], value)
                    if isinstance(value, (int, float)):
                        cell.value = value
                    elif isinstance(value, str):
                        cell.value = value
                    else:
                        try:
                            cell.value = json.dumps(value)
                        except Exception:
                            cell.value = str(value)
                    remapped_count += 1
                else:
                    log_append(text_area_placeholder, logs, f"[RULE3] Could not remap '{expr}'.")
                continue

            # If file exists, do strict evaluation on image
            if image_key.startswith("voicetest"):
                log_append(text_area_placeholder, logs, f"[RULE3] Strictly evaluating voice image '{image_key}'.")
                voice_eval = evaluate_voice_image(token, file_path, model_generic, text_area_placeholder, logs)
                if voice_eval and 'data' in voice_eval:
                    voice_test.setdefault(image_key, {}).update(voice_eval['data'])
                    nested_keys = keys[1:]
                    resolved_after = resolve_expression_with_vars(expr, {**allowed_vars, "voice_test": voice_test})
                    if resolved_after is not None:
                        set_nested_value_case_insensitive(voice_test, nested_keys, resolved_after)
                        if isinstance(resolved_after, (int, float)):
                            cell.value = resolved_after
                        elif isinstance(resolved_after, str):
                            cell.value = resolved_after
                        else:
                            try:
                                cell.value = json.dumps(resolved_after)
                            except Exception:
                                cell.value = str(resolved_after)
                        remapped_count += 1
                        continue
                log_append(text_area_placeholder, logs, f"[RULE3] Asking model for value of '{expr}' using 'voice_test' variable.")
                value = ask_model_for_expression_value(token, "voice_test", voice_test, expr, model_generic, text_area_placeholder, logs)
                if value is not None:
                    set_nested_value_case_insensitive(voice_test, keys[1:], value)
                    cell.value = value if not isinstance(value, dict) else json.dumps(value)
                    remapped_count += 1
                else:
                    log_append(text_area_placeholder, logs, f"[RULE3] Could not remap '{expr}' from voice image.")
                continue
            else:
                # generic speed/video
                log_append(text_area_placeholder, logs, f"[RULE3] Strictly evaluating generic image '{image_key}'.")
                gen_eval = evaluate_generic_image(token, file_path, model_generic, text_area_placeholder, logs)
                if gen_eval and 'data' in gen_eval:
                    pref = image_key.split('_')[0]
                    if pref == "alpha":
                        if gen_eval['image_type'] == 'speed_test':
                            alpha_speedtest.setdefault(image_key, {}).update(gen_eval['data'])
                        elif gen_eval['image_type'] == 'video_test':
                            alpha_video.setdefault(image_key, {}).update(gen_eval['data'])
                    elif pref == "beta":
                        if gen_eval['image_type'] == 'speed_test':
                            beta_speedtest.setdefault(image_key, {}).update(gen_eval['data'])
                        elif gen_eval['image_type'] == 'video_test':
                            beta_video.setdefault(image_key, {}).update(gen_eval['data'])
                    elif pref == "gamma":
                        if gen_eval['image_type'] == 'speed_test':
                            gamma_speedtest.setdefault(image_key, {}).update(gen_eval['data'])
                        elif gen_eval['image_type'] == 'video_test':
                            gamma_video.setdefault(image_key, {}).update(gen_eval['data'])

                    # attempt to resolve after updating
                    new_allowed = {
                        "alpha_service": alpha_service, "beta_service": beta_service, "gamma_service": gamma_service,
                        "alpha_speedtest": alpha_speedtest, "beta_speedtest": beta_speedtest, "gamma_speedtest": gamma_speedtest,
                        "alpha_video": alpha_video, "beta_video": beta_video, "gamma_video": gamma_video,
                        "voice_test": voice_test, "avearge": avearge
                    }
                    resolved_after = resolve_expression_with_vars(expr, new_allowed)
                    if resolved_after is not None:
                        nested_keys = key_pattern.findall(rest)
                        set_nested_value_case_insensitive(new_allowed[base_key], nested_keys, resolved_after)
                        allowed_vars = new_allowed
                        if isinstance(resolved_after, (int, float)):
                            cell.value = resolved_after
                        elif isinstance(resolved_after, str):
                            cell.value = resolved_after
                        else:
                            try:
                                cell.value = json.dumps(resolved_after)
                            except Exception:
                                cell.value = str(resolved_after)
                        remapped_count += 1
                        continue

                # if still not resolved, ask model using base var
                log_append(text_area_placeholder, logs, f"[RULE3] Asking model for value of '{expr}' using variable '{base_key}'.")
                value = ask_model_for_expression_value(token, base_key, allowed_vars[base_key], expr, model_generic, text_area_placeholder, logs)
                if value is not None:
                    nested_keys = key_pattern.findall(rest)
                    set_nested_value_case_insensitive(allowed_vars[base_key], nested_keys, value)
                    if isinstance(value, (int, float)):
                        cell.value = value
                    elif isinstance(value, str):
                        cell.value = value
                    else:
                        try:
                            cell.value = json.dumps(value)
                        except Exception:
                            cell.value = str(value)
                    remapped_count += 1
                else:
                    log_append(text_area_placeholder, logs, f"[RULE3] Could not remap '{expr}'. Left as NULL.")

        wb_r3.save(local_template)
        log_append(text_area_placeholder, logs, f"[RULE3] Remapping complete. Cells remapped: {remapped_count}. Workbook saved.")

    except Exception as e:
        log_append(text_area_placeholder, logs, f"[ERROR] Rule 3 remapping failed: {e}")

    # -------------------- PRINT FINAL OUTPUTS IN LOG --------------------
    log_append(text_area_placeholder, logs, "\n" + "="*50)
    log_append(text_area_placeholder, logs, "--- FINAL EXTRACTED AND STRUCTURED DATA (POST-EVAL & RULE2 & RULE3) ---")
    log_append(text_area_placeholder, logs, "="*50)

    def _pp(title, obj):
        try:
            s = json.dumps(obj, indent=2)
        except Exception:
            s = str(obj)
        log_append(text_area_placeholder, logs, f"\n{title}:")
        log_append(text_area_placeholder, logs, s)

    _pp("alpha_service", alpha_service)
    _pp("beta_service", beta_service)
    _pp("gamma_service", gamma_service)
    _pp("alpha_speedtest", alpha_speedtest)
    _pp("beta_speedtest", beta_speedtest)
    _pp("gamma_speedtest", gamma_speedtest)
    _pp("alpha_video", alpha_video)
    _pp("beta_video", beta_video)
    _pp("gamma_video", gamma_video)
    _pp("voice_test", voice_test)
    _pp("avearge", avearge)
    _pp("extract_text", extract_text)

    # return path to updated workbook (local_template)
    return local_template

# ----------------- Streamlit UI -----------------
def validate_api_key(token: str) -> Tuple[bool, str]:
    """
    Minimal token validation:
     - Quick heuristic: token should include 'apify_api'
     - Optionally: make a small call to the API to detect 401 vs success (this may consume credits).
    We'll perform a cheap HEAD/OPTIONS-style call if possible (but OpenRouter likely needs POST).
    For safety, use substring check and return True if passes; otherwise False.
    """
    if not token or "apify_api" not in token:
        return False, "API key does not look like an Apify token (missing 'apify_api')."
    # Optionally, you can attempt a tiny request here to validate, but it may consume credits.
    return True, "Token looks valid (format check)."

def main_ui():
    st.set_page_config(page_title="Advanced Cellular Template Processor", layout="wide")
    st.title("Advanced Cellular Template Processor")
    st.write("Upload templates and process images with the AI extractor. Set API key in the sidebar first.")

    # Sidebar: API key and validation
    st.sidebar.header("API Key & Settings")
    token_input = st.sidebar.text_input("Enter your Apify/OpenRouter API token", type="password", placeholder="apify_api_...")
    validate_btn = st.sidebar.button("Validate API key")

    # session logs
    if 'logs' not in st.session_state:
        st.session_state['logs'] = []
    # placeholder for logs
    log_placeholder = st.empty()
    log_placeholder.text_area("Logs", "\n".join(st.session_state['logs'][-1000:]), height=300)

    if validate_btn:
        ok, msg = validate_api_key(token_input)
        if ok:
            st.sidebar.success("API token accepted (format). You may proceed.")
            st.session_state['APIFY_TOKEN'] = token_input
            st.session_state['API_VALID'] = True
            st.session_state['logs'].append("[UI] API key stored in session.")
            log_placeholder.text_area("Logs", "\n".join(st.session_state['logs'][-1000:]), height=300)
        else:
            st.sidebar.error(f"Validation failed: {msg}")
            st.session_state['API_VALID'] = False

    # Only show upload UI if API validated
    if st.session_state.get('API_VALID', False):
        st.header("Upload & Process")
        uploaded_file = st.file_uploader("Upload template (.xlsx or .csv)", type=['xlsx', 'csv'], accept_multiple_files=False)
        st.write("If you uploaded a CSV, also upload images (or zip) below.")
        uploaded_images = st.file_uploader("Upload images (png/jpg/jpeg) or ZIP (multiple)", type=['png','jpg','jpeg','zip'], accept_multiple_files=True)

        # model selection (optional)
        model_service = st.selectbox("Model for SERVICE images", options=[MODEL_SERVICE_DEFAULT], index=0)
        model_generic = st.selectbox("Model for GENERIC images", options=[MODEL_GENERIC_DEFAULT], index=0)

        if uploaded_file:
            # write uploaded_file to temp and show preview + process button
            tmp_dir = "streamlit_temp"
            if os.path.exists(tmp_dir):
                shutil.rmtree(tmp_dir)
            os.makedirs(tmp_dir, exist_ok=True)
            saved_template_path = os.path.join(tmp_dir, uploaded_file.name)
            with open(saved_template_path, "wb") as f:
                f.write(uploaded_file.read())
            st.success(f"Saved uploaded file: {uploaded_file.name}")

            process_btn = st.button("Process file now")
            if process_btn:
                # start processing
                st.session_state['logs'].append("[UI] Starting processing...")
                log_placeholder.text_area("Logs", "\n".join(st.session_state['logs'][-1000:]), height=300)

                out = process_file_streamlit(user_file_path=saved_template_path,
                                            token=st.session_state['APIFY_TOKEN'],
                                            temp_dir=tmp_dir,
                                            image_uploads=uploaded_images,
                                            logs=st.session_state['logs'],
                                            text_area_placeholder=log_placeholder,
                                            model_service=model_service,
                                            model_generic=model_generic)
                if out:
                    st.success("Processing finished.")
                    with open(out, "rb") as f:
                        st.download_button("Download processed file", data=f, file_name=os.path.basename(out))
                else:
                    st.error("Processing failed — check logs.")
    else:
        st.info("Please validate an API key in the sidebar before uploading files.")

if __name__ == "__main__":
    main_ui()
